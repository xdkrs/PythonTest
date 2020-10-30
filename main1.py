#Python test Xogito - Cristian Vladu
#This code uses PEP 8 Style and Naming convetion

#Import used libraries
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
import re
import requests
import os

#Create folder for Assets to download if it does not exist
if not os.path.exists('Assets'):
    os.mkdir('Assets')

#Set styles for formatting
header_fill = PatternFill(start_color='AABBCC',
                          end_color='AABBCC',
                          fill_type='solid')

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

#Open provided file
wbook = openpyxl.load_workbook(filename="python practice test.xlsx")
wsheet = wbook.active

#Initialize the final vector
final = []
final.append(["Product Code", "Danger column", "Warning column"])

#Count the number of rows
nr_rows = len([row for row in wsheet if not all([cell.value == None for cell in row])])

#Loop through all the rows of the provided file
for i in range(2, nr_rows+1):
    #Check if Danger or Warning keywords found.
    val = wsheet.cell(row=i, column=7).value
    #Convert to lowercase for a better match
    val = val.lower()
    d = 'danger' in val
    w = 'warning' in val
    
    #If values 'danger' or 'warning' found, store the desired values in the final vector
    if d:
        final.append([wsheet.cell(row=i+1, column=6).value, wsheet.cell(row=i, column=7).value, ''])

    elif w:
        final.append([wsheet.cell(row=i+1, column=6).value, '', wsheet.cell(row=i, column=7).value])
    
    #Download resources if 'danger' or 'warning' found
    if d or w :
        response = requests.get(wsheet.cell(row=i, column=16).value)
        name = (str(wsheet.cell(row=i+1, column=6).value) + '.jpg') if os.path.splitext(wsheet.cell(row=i, column=16).value)[1] == '.jpg' else (str(wsheet.cell(row=i+1, column=6).value) + '.html')
        with open(os.path.join('Assets', name), 'wb') as f:
            f.write(response.content)

#Open new excel workbook
wbook_result = openpyxl.Workbook()
wsheet_result = wbook_result.active

#Initialize current row, for formating
curr_row = 1

#For each value in the final vector, store it line by line in the new excel file
for row in final:
    wsheet_result.append(row)
    wsheet_result.cell(row=curr_row,column=1).border= thin_border
    wsheet_result.cell(row=curr_row,column=2).border= thin_border
    wsheet_result.cell(row=curr_row,column=3).border= thin_border
    curr_row += 1 #Increment current row

#Set fill color in the export excel file
wsheet_result['A1'].fill = header_fill
wsheet_result['B1'].fill = header_fill
wsheet_result['C1'].fill = header_fill

#Save and close workbook
wbook_result.save('final.xlsx')